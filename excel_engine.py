# auditoria_engine.py
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import COLUMNAS_DIAN_VS_CONT, OUTPUT_SHEETS_TO_HIDE, ORDEN, RED_FILL_COLOR, CARACTERES_ESPECIALES
from errors import ErrorSistema, ErrorUsuario, HojaNoEncontradaError, logger


class ConciliadorAuditoria:
    def __init__(self, file_path, sheet_names, seriales_iva=None, seriales_base=None,
                 seriales_base2=None, seriales_ret=None, progress_callback=None):
        self.file_path = Path(file_path)
        self.sheet_names = sheet_names
        self.progress_callback = progress_callback or (lambda mensaje: None)
        self.seriales_iva = seriales_iva if seriales_iva is not None else []
        self.seriales_base = seriales_base if seriales_base is not None else []
        self.seriales_base2 = seriales_base2 if seriales_base2 is not None else []
        self.seriales_ret = seriales_ret if seriales_ret is not None else []

    def _reportar(self, mensaje):
        logger.info(mensaje)
        self.progress_callback(mensaje)

    def _validar_hojas_existen(self, hojas_excel):
        for clave in ('principal', 'aud_comp'):
            nombre = self.sheet_names.get(clave)
            if not nombre:
                raise ErrorUsuario(f"No se especificó el nombre de la hoja '{clave}'.")
            if nombre not in hojas_excel:
                raise HojaNoEncontradaError(nombre, hojas_excel)

    @staticmethod
    def _limpiar_encabezados(df, mayus=False):
        cols = df.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip()
        df.columns = cols.str.upper() if mayus else cols
        return df
    
    @staticmethod
    def _shift_table_ref(ref, pivot_col, delta):
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if pivot_col <= max_col:
            max_col += delta
        if delta > 0 and pivot_col <= min_col:
            min_col += delta
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    @staticmethod
    def _cargar_hoja_con_encabezado_variable(path, sheet):
        try:
            df_temp = pd.read_excel(path, sheet_name=sheet, header=None, nrows=10)
            header_idx = 0
            for idx, row in df_temp.iterrows():
                if row.dropna().count() > 2:
                    header_idx = idx
                    break
            return pd.read_excel(path, sheet_name=sheet, header=header_idx)
        except Exception as e:
            raise ErrorSistema(f"Error leyendo la hoja '{sheet}' con encabezado variable: {e}") from e

    def ejecutar(self):
        if not self.file_path.exists():
            raise ErrorUsuario(f"El archivo no existe: {self.file_path}")

        try:
            xls = pd.ExcelFile(self.file_path)
        except Exception as e:
            raise ErrorSistema(f"No se pudo abrir el archivo Excel: {e}") from e

        hojas_excel = xls.sheet_names
        self._validar_hojas_existen(hojas_excel)

        self._reportar("Cargando hoja de Token procesado...")
        try:
            df_full = pd.read_excel(self.file_path, sheet_name=self.sheet_names['principal'])
            df_full = self._limpiar_encabezados(df_full)
        except Exception as e:
            raise ErrorSistema(f"Error leyendo la hoja principal: {e}") from e

        for col in ('CONCEPTO', 'TERCERO', 'TIPO', 'BASE', 'Num.Ext'):
            if col not in df_full.columns:
                raise ErrorUsuario(
                    "La hoja de Token no está formateada. Ejecuta primero 'Formatear Token' antes de la auditoría."
                )

        self._reportar("Cargando hoja de auditoría de comprobantes...")
        df_aud_comp = self._cargar_hoja_con_encabezado_variable(self.file_path, self.sheet_names['aud_comp'])
        df_aud_comp = self._limpiar_encabezados(df_aud_comp)
        df_aud_comp = df_aud_comp.dropna(subset=df_aud_comp.columns[:6], how='all').copy()

        nombre_auto = self.sheet_names.get('autorretenedores')
        if nombre_auto and nombre_auto in hojas_excel:
            try:
                df_autoretenedores = pd.read_excel(self.file_path, sheet_name=nombre_auto)
                df_autoretenedores = self._limpiar_encabezados(df_autoretenedores, mayus=True)
            except Exception as e:
                logger.warning(f"No se pudo leer la hoja de autorretenedores '{nombre_auto}': {e}")
                df_autoretenedores = pd.DataFrame(columns=['NIT', 'COMENTARIO'])
        else:
            df_autoretenedores = pd.DataFrame(columns=['NIT', 'COMENTARIO'])

        self._reportar("Filtrando registros y mapeando motivos de diferencias...")
        
        # --- EXCLUSIÓN DEFINITIVA DE EMITIDOS ---
        col_grupo = next((c for c in df_full.columns if str(c).strip().lower() == 'grupo'), None)
        if col_grupo:
            mask_emi = df_full[col_grupo].astype(str).str.contains('emitido', case=False, na=False)
            df_full = df_full[~mask_emi].reset_index(drop=True)

        # --- 1. MAPEO DE RAZONES DE EXCLUSIÓN PARA DIAN ---
        df_full['MOTIVO_ALERTA'] = ''
        
        es_personales_full = df_full['TIPO'].astype(str).str.strip().str.upper() == 'PERSONALES'
        df_full.loc[es_personales_full, 'MOTIVO_ALERTA'] = 'DIAN - Gasto Personal'

        mask_ignorar = pd.Series(False, index=df_full.index)
        
        col_tipo_doc = next((c for c in df_full.columns if str(c).strip().lower() == 'tipo de documento'), None)
        if col_tipo_doc:
            mask_app = df_full[col_tipo_doc].astype(str).str.contains('application response', case=False, na=False)
            mask_ignorar |= mask_app
            df_full.loc[mask_app & (df_full['MOTIVO_ALERTA'] == ''), 'MOTIVO_ALERTA'] = 'DIAN - Application Response (Evento)'
            
        mask_dian_sin_num = df_full['Num.Ext'].fillna('').astype(str).str.strip() == ''
        df_full.loc[mask_dian_sin_num & (df_full['MOTIVO_ALERTA'] == ''), 'MOTIVO_ALERTA'] = 'DIAN - Sin Num.Ext'

        df_proc = df_full[~mask_ignorar].copy()
        es_personales_proc = es_personales_full[~mask_ignorar]

        col_concepto_res = 'TIPO-DETALLE' if 'TIPO-DETALLE' in df_proc.columns else 'CONCEPTO'
        df_resultado = df_proc[~es_personales_proc][['Prefijo', 'BASE', 'Num.Ext', col_concepto_res]]
        if col_concepto_res != 'TIPO-DETALLE':
            df_resultado = df_resultado.rename(columns={col_concepto_res: 'TIPO-DETALLE'})

        for col in ORDEN:
            if col not in df_full.columns:
                df_full[col] = ''
        
        df_auditoria = df_full[ORDEN + ['MOTIVO_ALERTA']].copy()

        self._reportar("Procesando hoja de auditoría de comprobantes...")
        df_res_auditoria = self._procesar_aud_comp(
            df_aud_comp, self.seriales_iva, self.seriales_base, self.seriales_base2, self.seriales_ret
        )

        self._reportar("Unificando y conciliando DIAN vs Contabilidad...")
        df_dian_vs_cont, parejas_incompletas, dif_base, dif_iva = self._unificar_dian_vs_cont(
            df_auditoria, df_aud_comp, df_res_auditoria, df_autoretenedores
        )

        self._reportar("Escribiendo resultados en el archivo Excel...")
        self._escribir_excel(
            df_resultado=df_resultado,
            df_auditoria=df_auditoria.drop(columns=['MOTIVO_ALERTA'], errors='ignore'),
            df_res_auditoria=df_res_auditoria,
            df_dian_vs_cont=df_dian_vs_cont,
            parejas_incompletas=parejas_incompletas,
            df_autoretenedores=df_autoretenedores,
            dif_base=dif_base,
            dif_iva=dif_iva
        )

        self._reportar("Proceso completado con éxito.")
        return {
            "filas_procesadas": len(df_full),
            "filas_personales": int(es_personales_full.sum()),
            "filas_sin_pareja": int(sum(parejas_incompletas)),
        }

    @staticmethod
    def _procesar_aud_comp(df_aud_comp, seriales_iva, seriales_base, seriales_base2, seriales_ret=None):
        cols_iva = [col for col in seriales_iva if col in df_aud_comp.columns]
        cols_base = [col for col in seriales_base if col in df_aud_comp.columns]
        cols_base2 = [col for col in seriales_base2 if col in df_aud_comp.columns]

        df_res_auditoria = pd.DataFrame(index=df_aud_comp.index)

        df_res_auditoria['Num.Ext'] = (
            df_aud_comp['Num.Ext'].fillna('').astype(str) if 'Num.Ext' in df_aud_comp.columns else ''
        )
        df_res_auditoria['IVA'] = (
            df_aud_comp[cols_iva].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) * -1 if cols_iva else 0
        )
        df_res_auditoria['BASE'] = (
            df_aud_comp[cols_base].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) * -1 if cols_base else 0
        )
        df_res_auditoria['BASE_2'] = (
            df_aud_comp[cols_base2].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) * -1 if cols_base2 else 0
        )
        return df_res_auditoria[['Num.Ext', 'BASE', 'BASE_2', 'IVA']]

    @staticmethod
    def _unificar_dian_vs_cont(df_auditoria, df_aud_comp, df_res_auditoria, df_autoretenedores):
        df_dian_prep = df_auditoria.copy()
        df_dian_prep['Fecha'] = ''
        df_dian_prep['BASE_2'] = ''
        df_dian_prep['Prioridad_Fila'] = 1
        for col in COLUMNAS_DIAN_VS_CONT:
            if col not in df_dian_prep.columns:
                df_dian_prep[col] = ''

        df_cont_prep = pd.DataFrame()
        df_cont_prep['Tipo de documento'] = df_aud_comp['Tipo'] if 'Tipo' in df_aud_comp.columns else ''

        if 'Número' in df_aud_comp.columns:
            numero_limpio = (
                df_aud_comp['Número'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            )
            df_cont_prep['CUFE/CUDE'] = numero_limpio.apply(lambda x: x.zfill(10) if x.isdigit() else x)
        else:
            df_cont_prep['CUFE/CUDE'] = ''

        # --- LIMPIEZA DE CARACTERES ESPECIALES EN Num.Ext ---
        patron_regex = f"[{re.escape(''.join(CARACTERES_ESPECIALES))}]"

        df_cont_prep['Num.Ext'] = df_aud_comp['Num.Ext'] if 'Num.Ext' in df_aud_comp.columns else ''
        df_cont_prep['Num.Ext_Clean'] = (
            df_cont_prep['Num.Ext']
            .fillna('')
            .astype(str)
            .str.replace(patron_regex, '', regex=True)
            .str.strip()
        )
        
        df_cont_prep['MOTIVO_ALERTA'] = np.where(df_cont_prep['Num.Ext_Clean'] == '', 'CONT - Sin Num.Ext', '')

        df_cont_prep['Fecha'] = df_aud_comp['Fecha'] if 'Fecha' in df_aud_comp.columns else ''
        df_cont_prep['NIT Emisor'] = df_aud_comp['Nit/C.C.'] if 'Nit/C.C.' in df_aud_comp.columns else ''
        df_cont_prep['Nombre Emisor'] = df_aud_comp['Tercero'] if 'Tercero' in df_aud_comp.columns else ''
        df_cont_prep['BASE'] = df_res_auditoria['BASE']
        df_cont_prep['BASE_2'] = df_res_auditoria['BASE_2']
        df_cont_prep['IVA'] = df_res_auditoria['IVA']
        df_cont_prep['Prioridad_Fila'] = 2
        
        for col in COLUMNAS_DIAN_VS_CONT:
            if col not in df_cont_prep.columns:
                df_cont_prep[col] = ''

        df_dian_prep['Num.Ext_Clean'] = (
            df_dian_prep['Num.Ext']
            .fillna('')
            .astype(str)
            .str.replace(patron_regex, '', regex=True)
            .str.strip()
        )

        df_unificado = pd.concat([df_dian_prep, df_cont_prep], ignore_index=True)

        # --- FILTROS ADICIONALES: IGNORAR "TOTALES" Y "0000000000" ---
        col_nombre_emisor = 'Nombre Emisor' if 'Nombre Emisor' in df_unificado.columns else 'TERCERO'
        mask_totales = df_unificado[col_nombre_emisor].fillna('').astype(str).str.upper().str.contains('TOTALES')
        mask_ceros = df_unificado['Num.Ext_Clean'].fillna('').astype(str).str.contains('0000000000')
        
        df_unificado = df_unificado[~(mask_totales | mask_ceros)].reset_index(drop=True)
        
        # --- LÓGICA DE EMPAREJAMIENTO EXCLUSIVA PARA REGISTROS APTOS ---
        df_unificado['Conteo_Pareja'] = 1
        mask_matchable = (df_unificado['Num.Ext_Clean'] != '') & (df_unificado['MOTIVO_ALERTA'] == '')
        
        counts = df_unificado[mask_matchable].groupby('Num.Ext_Clean')['Num.Ext_Clean'].transform('count')
        df_unificado.loc[mask_matchable, 'Conteo_Pareja'] = counts

        mask_single = mask_matchable & (df_unificado['Conteo_Pareja'] == 1)
        is_dian = df_unificado['Prioridad_Fila'] == 1
        is_cont = df_unificado['Prioridad_Fila'] == 2
        
        df_unificado.loc[mask_single & is_dian, 'MOTIVO_ALERTA'] = 'DIAN - Sin Pareja en Contabilidad'
        df_unificado.loc[mask_single & is_cont, 'MOTIVO_ALERTA'] = 'CONT - Sin Pareja en DIAN'

        df_unificado['NIT_Emisor_Clean'] = (
            df_unificado['NIT Emisor'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        )
        df_unificado['NIT_Emisor_Group'] = df_unificado.groupby('Num.Ext_Clean')['NIT_Emisor_Clean'].transform('max')

        if not df_autoretenedores.empty:
            col_nit_auto = next((c for c in df_autoretenedores.columns if 'NIT' in c), None)
            col_coment_auto = next((c for c in df_autoretenedores.columns if 'COMENT' in c), None)

            if col_nit_auto and col_coment_auto:
                df_auto_clean = df_autoretenedores[[col_nit_auto, col_coment_auto]].copy()
                df_auto_clean['NIT_Clean'] = (
                    df_auto_clean[col_nit_auto].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                )
                df_auto_clean = df_auto_clean.drop_duplicates(subset=['NIT_Clean'])

                df_unificado = df_unificado.merge(
                    df_auto_clean[['NIT_Clean', col_coment_auto]],
                    left_on='NIT_Emisor_Group',
                    right_on='NIT_Clean',
                    how='left'
                )

                cond_comentario = (df_unificado['Prioridad_Fila'] == 2) & (df_unificado[col_coment_auto].notna())
                df_unificado['NIT Receptor'] = np.where(
                    cond_comentario, df_unificado[col_coment_auto], df_unificado['NIT Receptor']
                )
                df_unificado = df_unificado.drop(columns=['NIT_Clean', col_coment_auto], errors='ignore')

        df_unificado['Tipo_Doc_Clean'] = df_unificado['Tipo de documento'].fillna('').astype(str).str.strip().str.upper()
        df_unificado['Tipo_Doc_Group'] = np.where(df_unificado['Tipo_Doc_Clean'] == 'FC1', 1, 2)
        df_unificado['Tipo_Doc_Group'] = df_unificado.groupby('Num.Ext_Clean')['Tipo_Doc_Group'].transform('min')

        df_unificado['Es_Incompleta'] = df_unificado['MOTIVO_ALERTA'] != ''

        df_unificado['Sort_Incompleta_Tipo'] = np.where(df_unificado['Es_Incompleta'], df_unificado['Tipo_Doc_Clean'], '')
        df_unificado['Sort_Incompleta_Motivo'] = np.where(df_unificado['Es_Incompleta'], df_unificado['MOTIVO_ALERTA'], '')

        df_unificado = df_unificado.sort_values(
            by=['Es_Incompleta', 'Sort_Incompleta_Tipo', 'Sort_Incompleta_Motivo', 'Tipo_Doc_Group', 'NIT_Emisor_Group', 'Num.Ext_Clean', 'Prioridad_Fila'],
            ascending=[True, True, True, True, False, True, True]
        )

        # --- REVISIÓN ROBUSTA DE DIFERENCIAS EN PAREJAS ---
        base_val_numeric = pd.to_numeric(df_unificado['BASE'], errors='coerce').fillna(0)
        iva_val_numeric = pd.to_numeric(df_unificado['IVA'], errors='coerce').fillna(0)

        df_unificado['Temp_Base_Num'] = base_val_numeric
        df_unificado['Temp_IVA_Num'] = iva_val_numeric
        
        mask_pairs = df_unificado['Conteo_Pareja'] >= 2
        
        # Agrupamos por pareja y sumamos para evaluar la diferencia total de la pareja
        group_base_sum = df_unificado[mask_pairs].groupby('Num.Ext_Clean')['Temp_Base_Num'].transform('sum')
        group_iva_sum = df_unificado[mask_pairs].groupby('Num.Ext_Clean')['Temp_IVA_Num'].transform('sum')

        UMBRAL_TOLERANCIA = 50.0
        df_unificado['Diff_BASE'] = False
        df_unificado['Diff_IVA'] = False
        
        df_unificado.loc[mask_pairs, 'Diff_BASE'] = group_base_sum.abs() > UMBRAL_TOLERANCIA
        df_unificado.loc[mask_pairs, 'Diff_IVA'] = group_iva_sum.abs() > UMBRAL_TOLERANCIA

        df_unificado = df_unificado.drop(columns=['Temp_Base_Num', 'Temp_IVA_Num'])

        parejas_incompletas = df_unificado['Es_Incompleta'].tolist()
        dif_base_list = df_unificado['Diff_BASE'].tolist()
        dif_iva_list = df_unificado['Diff_IVA'].tolist()

        columnas_finales = list(COLUMNAS_DIAN_VS_CONT)
        if 'MOTIVO_ALERTA' not in columnas_finales:
            columnas_finales.append('MOTIVO_ALERTA')

        df_dian_vs_cont = df_unificado[columnas_finales]
        return df_dian_vs_cont, parejas_incompletas, dif_base_list, dif_iva_list

    def _escribir_excel(
            self, df_resultado, df_auditoria, df_res_auditoria, df_dian_vs_cont,
            parejas_incompletas, df_autoretenedores=None, dif_base=None, dif_iva=None
        ):
        if df_autoretenedores is None:
            df_autoretenedores = pd.DataFrame(columns=['NIT', 'COMENTARIO'])
        if dif_base is None: dif_base = []
        if dif_iva is None: dif_iva = []

        try:
            engine_kwargs = {'keep_vba': True} if self.file_path.suffix.lower() == '.xlsm' else {}
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace',
                                 engine_kwargs=engine_kwargs) as writer:
                nombre_aud_comp = self.sheet_names['aud_comp']

                df_resultado.to_excel(writer, index=False, sheet_name='Resultados')
                df_auditoria.to_excel(writer, index=False, sheet_name='auditoria')
                df_dian_vs_cont.to_excel(writer, index=False, sheet_name='DIAN VS CONT')

                wb = writer.book
                sheet_resultados = writer.sheets['Resultados']
                sheet_auditoria = writer.sheets['auditoria']
                sheet_dian_vs_cont = writer.sheets['DIAN VS CONT']

                # --- INYECCIÓN Y ORDENAMIENTO EN HOJA ORIGINAL AUD-COMP ---
                sheet_aud_comp_orig = wb[nombre_aud_comp]
                header_row = 0
                col_tercero = 0
                col_nit = 0
                col_tipo = 0

                for r in range(1, 20):
                    for c in range(1, sheet_aud_comp_orig.max_column + 1):
                        val = str(sheet_aud_comp_orig.cell(row=r, column=c).value).strip().upper()
                        if val == 'TERCERO': col_tercero = c
                        elif val in ['NIT/C.C.', 'NIT']: col_nit = c
                        elif val == 'TIPO': col_tipo = c

                    if col_tercero > 0:
                        header_row = r
                        break

                if header_row > 0 and col_tercero > 0:
                    num_injected_found = 0
                    for i in range(1, 5):
                        val = str(sheet_aud_comp_orig.cell(row=header_row, column=col_tercero + i).value).strip().upper()
                        if val in ['AUTORRETENCION', 'BASE', 'BASE_2', 'IVA']:
                            num_injected_found += 1
                        else:
                            break
                    
                    if num_injected_found > 0:
                        sheet_aud_comp_orig.delete_cols(col_tercero + 1, amount=num_injected_found)

                    dict_auto = {}
                    col_nit_auto = next((c for c in df_autoretenedores.columns if 'NIT' in str(c).upper()), None)
                    col_coment_auto = next((c for c in df_autoretenedores.columns if 'COMENT' in str(c).upper()), None)

                    if col_nit_auto and col_coment_auto:
                        for _, r in df_autoretenedores.iterrows():
                            nit_str = str(r[col_nit_auto]).replace('.0', '').strip()
                            if nit_str and str(nit_str).lower() != 'nan':
                                dict_auto[nit_str] = str(r[col_coment_auto]).strip()

                    row_data_list = []
                    df_res_records = df_res_auditoria.to_dict('records')
                    max_row_orig = sheet_aud_comp_orig.max_row
                    max_col_orig = sheet_aud_comp_orig.max_column

                    for idx, excel_row_idx in enumerate(range(header_row + 1, max_row_orig + 1)):
                        cell_values = [sheet_aud_comp_orig.cell(row=excel_row_idx, column=c).value for c in range(1, max_col_orig + 1)]
                        calc_data = df_res_records[idx] if idx < len(df_res_records) else {'BASE': 0, 'BASE_2': 0, 'IVA': 0}

                        nit_val = str(cell_values[col_nit - 1]).replace('.0', '').strip() if col_nit > 0 else ''
                        comentario_auto = dict_auto.get(nit_val, '')

                        tipo_val = str(cell_values[col_tipo - 1]).strip().upper() if col_tipo > 0 else ''
                        tercero_val = str(cell_values[col_tercero - 1]).strip().upper() if col_tercero > 0 else ''

                        row_data_list.append({
                            'original_values': cell_values,
                            'base': calc_data.get('BASE', 0),
                            'base_2': calc_data.get('BASE_2', 0),
                            'iva': calc_data.get('IVA', 0),
                            'auto': comentario_auto,
                            'tipo': tipo_val,
                            'tercero': tercero_val
                        })

                    max_base_por_tercero = {}
                    for row in row_data_list:
                        t_str = str(row.get('tercero', '')).strip().upper()
                        try:
                            b_val = float(row.get('base', 0))
                        except (ValueError, TypeError):
                            b_val = 0.0
                        
                        if t_str not in max_base_por_tercero or b_val > max_base_por_tercero[t_str]:
                            max_base_por_tercero[t_str] = b_val

                    def sort_key(row):
                        auto_str = str(row.get('auto', '')).upper().strip()
                        tipo_str = str(row.get('tipo', '')).upper().strip()
                        tercero_str = str(row.get('tercero', '')).strip().upper()
                        
                        # --- NUEVO MANEJO DE PRIORIDADES SEGÚN REQUERIMIENTO ---
                        if 'AUTORRETENEDOR' in auto_str:
                            prioridad = 1
                        elif auto_str != '' or 'REGIMEN SIMPLE' in tipo_str:
                            prioridad = 2 # Si tiene algún dato en auto_str que no sea autorretenedor
                        else:
                            prioridad = 3
                            
                        try:
                            base_val = float(row.get('base', 0))
                        except (ValueError, TypeError):
                            base_val = 0.0
                            
                        max_base_grupo = max_base_por_tercero.get(tercero_str, 0.0)
                        
                        return (prioridad, -max_base_grupo, tercero_str, base_val)

                    row_data_list.sort(key=sort_key)

                    idx_insert = col_tercero + 1
                    sheet_aud_comp_orig.insert_cols(idx_insert, amount=4)

                    c_AUTORRETENCION = sheet_aud_comp_orig.cell(row=header_row, column=idx_insert)
                    c_AUTORRETENCION.value = 'AUTORRETENCION'
                    c_AUTORRETENCION.fill = PatternFill(start_color="cc99ff", end_color="cc99ff", fill_type="solid")

                    c_base = sheet_aud_comp_orig.cell(row=header_row, column=idx_insert + 1)
                    c_base.value = 'BASE'
                    c_base.fill = PatternFill(start_color="339966", end_color="339966", fill_type="solid")

                    c_base2 = sheet_aud_comp_orig.cell(row=header_row, column=idx_insert + 2)
                    c_base2.value = 'BASE_2'
                    c_base2.fill = PatternFill(start_color="33EAEA", end_color="33EAEA", fill_type="solid") 

                    c_iva = sheet_aud_comp_orig.cell(row=header_row, column=idx_insert + 3)
                    c_iva.value = 'IVA'
                    c_iva.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

                    for i, row_dict in enumerate(row_data_list):
                        excel_row_idx = header_row + 1 + i
                        orig_vals = row_dict['original_values']

                        for c in range(1, idx_insert):
                            sheet_aud_comp_orig.cell(row=excel_row_idx, column=c).value = orig_vals[c-1]

                        cell_autorretencion = sheet_aud_comp_orig.cell(row=excel_row_idx, column=idx_insert)
                        cell_autorretencion.value = row_dict['auto']

                        c_b = sheet_aud_comp_orig.cell(row=excel_row_idx, column=idx_insert + 1)
                        c_b.value = row_dict['base']
                        c_b.number_format = '#,##0.00'

                        c_b2 = sheet_aud_comp_orig.cell(row=excel_row_idx, column=idx_insert + 2)
                        c_b2.value = row_dict['base_2']
                        c_b2.number_format = '#,##0.00'

                        c_i = sheet_aud_comp_orig.cell(row=excel_row_idx, column=idx_insert + 3)
                        c_i.value = row_dict['iva']
                        c_i.number_format = '#,##0.00'

                        for c in range(idx_insert, len(orig_vals) + 1):
                            sheet_aud_comp_orig.cell(row=excel_row_idx, column=c + 4).value = orig_vals[c-1]

                for sheet_target, df_target in [
                    (sheet_resultados, df_resultado),
                    (sheet_auditoria, df_auditoria),
                    (sheet_dian_vs_cont, df_dian_vs_cont)
                ]:
                    for col_name in ['BASE', 'BASE_2', 'IVA', 'Total']:
                        if col_name in df_target.columns:
                            col_idx = df_target.columns.get_loc(col_name) + 1
                            for row in range(2, len(df_target) + 2):
                                sheet_target.cell(row=row, column=col_idx).number_format = '#,##0.00'

                # --- 3. ESTILOS VISUALES Y CORRECCIÓN DE COLORES EN DIAN VS CONT ---
                max_row = len(df_dian_vs_cont) + 1
                max_col = len(df_dian_vs_cont.columns)

                sheet_dian_vs_cont.freeze_panes = 'A2'

                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )

                for col_idx in range(1, max_col + 1):
                    cell = sheet_dian_vs_cont.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border

                fill_vacio = PatternFill(start_color="A6B7F5", end_color="A6B7F5", fill_type="solid")
                fill_llenado = PatternFill(start_color="627FF0", end_color="627FF0", fill_type="solid")
                fill_red_alert = PatternFill(start_color=RED_FILL_COLOR, end_color=RED_FILL_COLOR, fill_type="solid")

                COLOR_MOTIVOS = {
                    'DIAN - Gasto Personal': 'FCE4D6',                
                    'DIAN - Application Response (Evento)': 'D9D9D9', 
                    'DIAN - Sin Num.Ext': 'F2DCDB',                   
                    'CONT - Sin Num.Ext': 'F2DCDB',                   
                    'DIAN - Sin Pareja en Contabilidad': 'FFF2CC',    
                    'CONT - Sin Pareja en DIAN': 'E2EFDA',            
                }

                col_grupo_idx = df_dian_vs_cont.columns.get_loc('Grupo') + 1 if 'Grupo' in df_dian_vs_cont.columns else None
                col_motivo_idx = df_dian_vs_cont.columns.get_loc('MOTIVO_ALERTA') + 1 if 'MOTIVO_ALERTA' in df_dian_vs_cont.columns else None
                
                col_base_idx_dian = df_dian_vs_cont.columns.get_loc('BASE') + 1 if 'BASE' in df_dian_vs_cont.columns else None
                col_iva_idx_dian = df_dian_vs_cont.columns.get_loc('IVA') + 1 if 'IVA' in df_dian_vs_cont.columns else None

                for row_idx in range(2, max_row + 1):
                    es_vacio = True
                    if col_grupo_idx:
                        val_grupo = sheet_dian_vs_cont.cell(row=row_idx, column=col_grupo_idx).value
                        if val_grupo is not None and str(val_grupo).strip() != '':
                            es_vacio = False

                    motivo = sheet_dian_vs_cont.cell(row=row_idx, column=col_motivo_idx).value if col_motivo_idx else ''
                    
                    # Determinamos el color BASE de la fila
                    if motivo and str(motivo).strip() != '':
                        motivo_str = str(motivo).strip()
                        color_hex = COLOR_MOTIVOS.get(motivo_str, 'FFC7CE') 
                        fill_base = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    else:
                        fill_base = fill_vacio if es_vacio else fill_llenado

                    # Paso A: Aplicamos el color base a toda la fila
                    for col_idx in range(1, max_col + 1):
                        cell = sheet_dian_vs_cont.cell(row=row_idx, column=col_idx)
                        cell.fill = fill_base
                        cell.font = Font(name="Arial", size=10)
                        cell.border = thin_border
                        
                    # Paso B: Sobreescribimos de rojo SÓLO si hay diferencias de valor (Aplica a TODA la pareja)
                    # CORRECCIÓN DE NOMBRES AQUÍ
                    es_dif_base = dif_base[row_idx - 2]
                    es_dif_iva = dif_iva[row_idx - 2]
                    
                    if es_dif_base and col_base_idx_dian:
                        sheet_dian_vs_cont.cell(row=row_idx, column=col_base_idx_dian).fill = fill_red_alert
                    if es_dif_iva and col_iva_idx_dian:
                        sheet_dian_vs_cont.cell(row=row_idx, column=col_iva_idx_dian).fill = fill_red_alert

                for col in sheet_dian_vs_cont.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            pass
                    sheet_dian_vs_cont.column_dimensions[col_letter].width = max(max_len + 3, 12)

                for sheetname in OUTPUT_SHEETS_TO_HIDE:
                    if sheetname in wb.sheetnames:
                        wb[sheetname].sheet_state = 'hidden'

                if not any(ws.sheet_state == 'visible' for ws in wb.worksheets):
                    wb.worksheets[0].sheet_state = 'visible'

        except (ErrorUsuario, ErrorSistema):
            raise
        except Exception as e:
            raise ErrorSistema(f"Error escribiendo resultados en el archivo Excel: {e}") from e