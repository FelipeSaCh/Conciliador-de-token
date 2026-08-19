import pandas as pd
import numpy as np
import re

from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Ruta del archivo de Excel
file_path = r"C:\Users\USUARIO\Desktop\Nuevo Hoja de cálculo de Microsoft Excel.xlsx"

ORDEN = ['Tipo de documento', 'CUFE/CUDE', 'Folio', 'Prefijo', 'Num.Ext', 'Divisa', 'Forma de Pago', 
         'Medio de Pago', 'Fecha Emisión', 'Fecha Recepción', 'NIT Emisor', 'Nombre Emisor', 
         'NIT Receptor', 'Nombre Receptor', 'BASE', 'IVA', 'Total', 'Estado', 'Grupo']

COLUMNAS_DIAN_VS_CONT = [
    'Tipo de documento', 'CUFE/CUDE', 'Folio', 'Prefijo', 'Num.Ext', 'Divisa', 'Fecha', 
    'Forma de Pago', 'Medio de Pago', 'Fecha Emisión', 'Fecha Recepción', 'NIT Emisor', 
    'Nombre Emisor', 'NIT Receptor', 'Nombre Receptor', 'BASE', 'IVA', 'Total', 'Estado', 'Grupo'
]

# 1. FUNCIÓN PARA DETECTAR Y CARGAR TABLAS CON ENCABEZADOS DESPLAZADOS
def cargar_hoja_con_encabezado_variable(path, sheet):
    try:
        df_temp = pd.read_excel(path, sheet_name=sheet, header=None, nrows=10)
        header_idx = 0
        for idx, row in df_temp.iterrows():
            if row.dropna().count() > 2:
                header_idx = idx
                break
        return pd.read_excel(path, sheet_name=sheet, header=header_idx)
    except Exception:
        return pd.DataFrame()

# 2. CARGAR HOJAS DEL EXCEL
df_principal = pd.read_excel(file_path, sheet_name='Sheet1')
df_conta = pd.read_excel(file_path, sheet_name='contabilidad')
df_tercer = pd.read_excel(file_path, sheet_name='TERCEROS')
df_aud_comp = cargar_hoja_con_encabezado_variable(file_path, 'AUD-COMP')

try:
    df_autoretenedores = pd.read_excel(file_path, sheet_name='AUTORRETENEDORES')
    df_autoretenedores.columns = df_autoretenedores.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip().str.upper()
except Exception:
    df_autoretenedores = pd.DataFrame(columns=['NIT', 'COMENTARIO'])

# Limpieza básica de columnas
df_conta.columns = df_conta.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip().str.upper()
df_principal.columns = df_principal.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip()
df_tercer.columns = df_tercer.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip().str.upper()
df_aud_comp.columns = df_aud_comp.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip()

df_aud_comp = df_aud_comp.dropna(subset=df_aud_comp.columns[:6], how='all').copy()

# 3. TRATAMIENTO Y CRUCE DE DATOS
df_principal['NIT_Emisor_clean'] = df_principal['NIT Emisor'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
df_conta['NIT_clean'] = df_conta['NIT'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
df_tercer['NIT_clean'] = df_tercer['NIT'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

df_conta_unica = df_conta.drop_duplicates(subset=['NIT_clean'])

cols_a_traer = ['NIT_clean']
if 'TIPO-DETALLE' in df_conta_unica.columns:
    cols_a_traer.append('TIPO-DETALLE')
if 'TIPO' in df_conta_unica.columns:
    cols_a_traer.append('TIPO')

df = df_principal.merge(
    df_conta_unica[cols_a_traer],
    left_on='NIT_Emisor_clean',
    right_on='NIT_clean',
    how='left'
)

df['TIPO-DETALLE'] = df['TIPO-DETALLE'].fillna('')
df['TIPO'] = df['TIPO'].fillna('')

df['CONCEPTO'] = np.where(
    df['TIPO-DETALLE'].astype(str).str.strip() == '',
    'CONCEPTO NO CREADO',
    df['TIPO-DETALLE'].astype(str).str.strip()
)

nits_terceros_unicos = set(df_tercer['NIT_clean'].unique())
df['TERCERO'] = np.where(
    df['NIT_Emisor_clean'].isin(nits_terceros_unicos) & (df['NIT_Emisor_clean'] != ''),
    'CREADO',
    'NO CREADO'
)

# 4. CONCATENACIÓN Y CÁLCULOS
prefijo_texto = df['Prefijo'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
folio_texto = df['Folio'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
num_ext_base = prefijo_texto + folio_texto

prefijo_esta_vacio = df['Prefijo'].isna() | (df['Prefijo'].astype(str).str.strip() == '')

df['Num.Ext'] = np.where(
    prefijo_esta_vacio,
    num_ext_base.str.zfill(10),
    num_ext_base
)

df['BASE'] = np.where(df['Total'] == 0, 0, (df['Total'] - df['IVA']) * -1)

# 5. PREPARAR HOJA RESULTADOS Y AUDITORIA
es_PERSONALES = df['TIPO'].astype(str).str.strip().str.upper() == 'PERSONALES'

df_resultado = df[~es_PERSONALES][['Prefijo', 'BASE', 'Num.Ext', 'TIPO-DETALLE']]

for col in ORDEN:
    if col not in df.columns:
        df[col] = ''

df_auditoria = df[~es_PERSONALES][ORDEN].copy()

# 6. PROCESAR HOJA AUD-COMP
seriales_iva = ['240811019', '240816019', '135530019', '531520002']
seriales_base = ['620501001', '620501019', '622501020', '52', '5305', '1540']

cols_iva = [col for col in df_aud_comp.columns if any(re.search(rf'\b{s}', str(col)) for s in seriales_iva)]
cols_base = [col for col in df_aud_comp.columns if any(re.search(rf'\b{s}', str(col)) for s in seriales_base)]

df_res_auditoria = pd.DataFrame()

df_res_auditoria['Num.Ext'] = df_aud_comp['Num.Ext'].fillna('').astype(str) if 'Num.Ext' in df_aud_comp.columns else ''
df_res_auditoria['IVA'] = df_aud_comp[cols_iva].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) if cols_iva else 0
df_res_auditoria['BASE'] = df_aud_comp[cols_base].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1) * -1 if cols_base else 0

df_res_auditoria = df_res_auditoria[['Num.Ext', 'BASE', 'IVA']]

# 7. UNIFICAR Y ORDENAR EN "DIAN VS CONT"
df_dian_prep = df_auditoria.copy()
df_dian_prep['Fecha'] = ''
df_dian_prep['Prioridad_Fila'] = 1

for col in COLUMNAS_DIAN_VS_CONT:
    if col not in df_dian_prep.columns:
        df_dian_prep[col] = ''

df_cont_prep = pd.DataFrame()
df_cont_prep['Tipo de documento'] = df_aud_comp['Tipo'] if 'Tipo' in df_aud_comp.columns else ''

if 'Número' in df_aud_comp.columns:
    numero_limpio = df_aud_comp['Número'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    df_cont_prep['CUFE/CUDE'] = numero_limpio.apply(lambda x: x.zfill(10) if x.isdigit() else x)
else:
    df_cont_prep['CUFE/CUDE'] = ''

df_cont_prep['Num.Ext'] = df_aud_comp['Num.Ext'] if 'Num.Ext' in df_aud_comp.columns else ''
df_cont_prep['Fecha'] = df_aud_comp['Fecha'] if 'Fecha' in df_aud_comp.columns else ''
df_cont_prep['NIT Emisor'] = df_aud_comp['Nit/C.C.'] if 'Nit/C.C.' in df_aud_comp.columns else ''
df_cont_prep['Nombre Emisor'] = df_aud_comp['Tercero'] if 'Tercero' in df_aud_comp.columns else ''
df_cont_prep['BASE'] = df_res_auditoria['BASE']
df_cont_prep['IVA'] = df_res_auditoria['IVA']
df_cont_prep['Prioridad_Fila'] = 2

for col in COLUMNAS_DIAN_VS_CONT:
    if col not in df_cont_prep.columns:
        df_cont_prep[col] = ''

df_dian_prep['Num.Ext_Clean'] = df_dian_prep['Num.Ext'].fillna('').astype(str).str.strip()
df_cont_prep['Num.Ext_Clean'] = df_cont_prep['Num.Ext'].fillna('').astype(str).str.strip()

df_unificado = pd.concat([df_dian_prep, df_cont_prep], ignore_index=True)
df_unificado['Conteo_Pareja'] = df_unificado.groupby('Num.Ext_Clean')['Num.Ext_Clean'].transform('count')

df_unificado['NIT_Emisor_Clean'] = df_unificado['NIT Emisor'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
df_unificado['NIT_Emisor_Group'] = df_unificado.groupby('Num.Ext_Clean')['NIT_Emisor_Clean'].transform('max')

# Búsqueda en Autorretenedores
if not df_autoretenedores.empty:
    col_nit_auto = next((c for c in df_autoretenedores.columns if 'NIT' in c), None)
    col_coment_auto = next((c for c in df_autoretenedores.columns if 'COMENT' in c), None)
    
    if col_nit_auto and col_coment_auto:
        df_auto_clean = df_autoretenedores[[col_nit_auto, col_coment_auto]].copy()
        df_auto_clean['NIT_Clean'] = df_auto_clean[col_nit_auto].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        df_auto_clean = df_auto_clean.drop_duplicates(subset=['NIT_Clean'])
        
        df_unificado = df_unificado.merge(
            df_auto_clean[['NIT_Clean', col_coment_auto]],
            left_on='NIT_Emisor_Group',
            right_on='NIT_Clean',
            how='left'
        )
        
        cond_comentario = (df_unificado['Prioridad_Fila'] == 2) & (df_unificado[col_coment_auto].notna())
        df_unificado['NIT Receptor'] = np.where(cond_comentario, df_unificado[col_coment_auto], df_unificado['NIT Receptor'])
        df_unificado = df_unificado.drop(columns=['NIT_Clean', col_coment_auto], errors='ignore')

# Ordenamiento
df_unificado['Tipo_Doc_Clean'] = df_unificado['Tipo de documento'].fillna('').astype(str).str.strip().str.upper()
df_unificado['Tipo_Doc_Group'] = np.where(df_unificado['Tipo_Doc_Clean'] == 'FC1', 1, 2)
df_unificado['Tipo_Doc_Group'] = df_unificado.groupby('Num.Ext_Clean')['Tipo_Doc_Group'].transform('min')

df_unificado = df_unificado.sort_values(
    by=['Tipo_Doc_Group', 'NIT_Emisor_Group', 'Num.Ext_Clean', 'Prioridad_Fila'],
    ascending=[True, False, True, True]
)

parejas_incompletas = (df_unificado['Conteo_Pareja'] == 1).tolist()
df_dian_vs_cont = df_unificado[COLUMNAS_DIAN_VS_CONT]

# 8. ESCRIBIR EN EXCEL, APLICAR FORMATOS, TABLA Y OCULTAR HOJAS GENERADAS
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    cols_sheet1_originales = [col for col in df_principal.columns if col != 'NIT_Emisor_clean']
    df_export_principal = df[cols_sheet1_originales + ['CONCEPTO', 'TERCERO']]
    
    df_export_principal.to_excel(writer, index=False, sheet_name='Sheet1')
    df_resultado.to_excel(writer, index=False, sheet_name='Resultados')
    df_auditoria.to_excel(writer, index=False, sheet_name='auditoria')
    df_res_auditoria.to_excel(writer, index=False, sheet_name='resultados-auditoria')
    df_dian_vs_cont.to_excel(writer, index=False, sheet_name='DIAN VS CONT')
    
    wb = writer.book
    sheet_principal = writer.sheets['Sheet1']
    sheet_resultados = writer.sheets['Resultados']
    sheet_auditoria = writer.sheets['auditoria']
    sheet_res_auditoria = writer.sheets['resultados-auditoria']
    sheet_dian_vs_cont = writer.sheets['DIAN VS CONT']
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Resaltado PERSONALES
    for row_idx, es_pers in enumerate(es_PERSONALES, start=2):
        if es_pers:
            for col_idx in range(1, len(df_export_principal.columns) + 1):
                sheet_principal.cell(row=row_idx, column=col_idx).fill = red_fill

    # Resaltado Filas sin pareja
    for row_idx, es_incompleta in enumerate(parejas_incompletas, start=2):
        if es_incompleta:
            for col_idx in range(1, len(COLUMNAS_DIAN_VS_CONT) + 1):
                sheet_dian_vs_cont.cell(row=row_idx, column=col_idx).fill = red_fill

    # Formato Moneda
    cop_format = '"$"#,##0'
    for sheet_target, df_target in [
        (sheet_resultados, df_resultado), 
        (sheet_auditoria, df_auditoria),
        (sheet_res_auditoria, df_res_auditoria),
        (sheet_dian_vs_cont, df_dian_vs_cont)
    ]:
        for col_name in ['BASE', 'IVA', 'Total']:
            if col_name in df_target.columns:
                col_idx = df_target.columns.get_loc(col_name) + 1
                for row in range(2, len(df_target) + 2):
                    sheet_target.cell(row=row, column=col_idx).number_format = cop_format

    # Crear Tabla de Excel en DIAN VS CONT
    max_row = len(df_dian_vs_cont) + 1
    max_col_letter = get_column_letter(len(COLUMNAS_DIAN_VS_CONT))
    tabla = Table(displayName="TablaDianVsCont", ref=f"A1:{max_col_letter}{max_row}")
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sheet_dian_vs_cont.add_table(tabla)

    # Autoajustar ancho de columnas en DIAN VS CONT
    for col in sheet_dian_vs_cont.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet_dian_vs_cont.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---------------------------------------------------------------------
    # OCULTAR ÚNICAMENTE LAS HOJAS INTERMEDIAS CREADAS POR EL SCRIPT
    # (Mantiene visibles las originales + 'DIAN VS CONT')
    # ---------------------------------------------------------------------
    hojas_a_ocultar = ['Resultados', 'auditoria', 'resultados-auditoria']
    for sheetname in wb.sheetnames:
        if sheetname in hojas_a_ocultar:
            wb[sheetname].sheet_state = 'hidden'

print("Proceso completado con éxito.")