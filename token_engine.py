# token_engine.py
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from config import RED_FILL_COLOR
from errors import ColumnaFaltanteError, ErrorSistema, ErrorUsuario, HojaNoEncontradaError, logger


class FormateadorToken:
    # Se añaden NIT Receptor y Grupo como requeridos para poder hacer la lógica condicional
    COLUMNAS_REQUERIDAS_PRINCIPAL = ['NIT Emisor', 'NIT Receptor', 'Grupo', 'Prefijo', 'Folio', 'Total', 'IVA']
    COLUMNAS_REQUERIDAS_CONTA = ['NIT']
    COLUMNAS_REQUERIDAS_TERCEROS = ['NIT']

    def __init__(self, file_path, sheet_names, progress_callback=None):
        self.file_path = Path(file_path)
        self.sheet_names = sheet_names
        self.progress_callback = progress_callback or (lambda mensaje: None)

    def _reportar(self, mensaje):
        logger.info(mensaje)
        self.progress_callback(mensaje)

    @staticmethod
    def _limpiar_encabezados(df, mayus=False):
        cols = df.columns.astype(str).str.replace(r'[\r\n\t]', '', regex=True).str.strip()
        df.columns = cols.str.upper() if mayus else cols
        return df

    @staticmethod
    def _leer_hoja_dinamica(file_path, sheet_name, keyword="NIT"):
        """
        Lee una hoja de Excel buscando dinámicamente en qué fila están los encabezados,
        basándose en la búsqueda de una columna clave.
        """
        # Leemos el excel sin asignar encabezado (header=None)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        header_idx = 0
        keyword_upper = keyword.upper()
        
        # Buscamos en las primeras 50 filas (por eficiencia)
        for idx in range(min(50, len(df))):
            # Limpiamos los valores de la fila actual para la comparación exacta
            fila_limpia = df.iloc[idx].astype(str).str.strip().str.upper()
            
            # Comprobamos si la palabra clave (ej. "NIT") existe en esta fila
            if keyword_upper in fila_limpia.values:
                header_idx = idx
                break
                
        # Asignamos la fila encontrada como los encabezados de las columnas
        df.columns = df.iloc[header_idx]
        
        # Nos quedamos solo con los datos reales (lo que está debajo del encabezado)
        df = df.iloc[header_idx + 1:].reset_index(drop=True)
        
        return df

    def _validar_columnas(self, df, columnas_requeridas, nombre_hoja):
        for col in columnas_requeridas:
            if col not in df.columns:
                raise ColumnaFaltanteError(nombre_hoja, col)

    def ejecutar(self):
        if not self.file_path.exists():
            raise ErrorUsuario(f"El archivo no existe: {self.file_path}")

        try:
            xls = pd.ExcelFile(self.file_path)
        except Exception as e:
            raise ErrorSistema(
                f"No se pudo abrir el archivo Excel: {e}"
            ) from e

        hojas_excel = xls.sheet_names
        nombre_principal = self.sheet_names.get("principal")
        if not nombre_principal:
            raise ErrorUsuario(
                "No se especificó el nombre de la hoja 'principal'."
            )
        if nombre_principal not in hojas_excel:
            raise HojaNoEncontradaError(nombre_principal, hojas_excel)

        self._reportar("Cargando hoja principal (Token)...")
        try:
            df_principal = pd.read_excel(
                self.file_path, sheet_name=nombre_principal
            )
            df_principal = self._limpiar_encabezados(df_principal)

            # --- LIMPIEZA DE FILAS COMPLETAMENTE VACÍAS ---
            df_principal = df_principal.dropna(how="all").reset_index(drop=True)

            self._validar_columnas(
                df_principal,
                self.COLUMNAS_REQUERIDAS_PRINCIPAL,
                nombre_principal,
            )
            
            # Guardamos las columnas originales exactas (sin contar las que calculamos nosotros)
            # Para garantizar la sobrescritura y no duplicarlas al exportar
            COLUMNAS_CALCULADAS = ["CONCEPTO", "TERCERO", "TIPO", "TIPO-DETALLE", "BASE", "Num.Ext"]
            cols_originales = [c for c in df_principal.columns if c not in COLUMNAS_CALCULADAS]
            
        except Exception as e:
            raise ErrorSistema(
                f"Error leyendo la hoja principal: {e}"
            ) from e

# -------------- CARGA DE CONTABILIDAD --------------
        nombre_conta = self.sheet_names.get("contabilidad")
        if nombre_conta and nombre_conta in hojas_excel:
            try:
                # Usamos el lector dinámico buscando la columna 'NIT'
                df_conta = self._leer_hoja_dinamica(self.file_path, nombre_conta, keyword="NIT")
                df_conta = self._limpiar_encabezados(df_conta, mayus=True)
                
                self._validar_columnas(
                    df_conta, self.COLUMNAS_REQUERIDAS_CONTA, nombre_conta
                )
            except Exception as e:
                logger.warning(
                    f"Error al procesar hoja de contabilidad '{nombre_conta}': {e}"
                )
                df_conta = pd.DataFrame(columns=self.COLUMNAS_REQUERIDAS_CONTA)
        else:
            df_conta = pd.DataFrame(columns=self.COLUMNAS_REQUERIDAS_CONTA)

        # -------------- CARGA DE TERCEROS --------------
        nombre_tercer = self.sheet_names.get("terceros")
        if nombre_tercer and nombre_tercer in hojas_excel:
            try:
                # Usamos el lector dinámico buscando la columna 'NIT'
                df_tercer = self._leer_hoja_dinamica(self.file_path, nombre_tercer, keyword="NIT")
                df_tercer = self._limpiar_encabezados(df_tercer, mayus=True)
                
                self._validar_columnas(
                    df_tercer, self.COLUMNAS_REQUERIDAS_TERCEROS, nombre_tercer
                )
            except Exception as e:
                logger.warning(
                    f"Error al procesar hoja de terceros '{nombre_tercer}': {e}"
                )
                df_tercer = pd.DataFrame(
                    columns=self.COLUMNAS_REQUERIDAS_TERCEROS
                )
        else:
            df_tercer = pd.DataFrame(columns=self.COLUMNAS_REQUERIDAS_TERCEROS)

        self._reportar("Cruzando datos entre hojas...")
        df_full = self._cruzar_datos(df_principal, df_conta, df_tercer)

        self._reportar("Calculando bases, totales y consecutivos...")
        df_full, es_personales = self._calcular_campos(df_full)

        self._reportar(
            "Ordenando registros por Grupo, Tipo de Documento y Personales..."
        )
        
        cols_export = cols_originales + COLUMNAS_CALCULADAS
        df_export = df_full[cols_export].copy()

        # Guardar marca de personales para mantener sincronizada la mascara de pintado
        df_export["_ES_PERSONAL"] = es_personales.values

        # --- ORDENAMIENTO POR JERARQUÍA ---
        df_export = self._ordenar_jerarquico(df_export)

        es_personales_ordenado = df_export["_ES_PERSONAL"]
        df_export = df_export.drop(columns="_ES_PERSONAL")

        self._reportar("Escribiendo hoja Token formateada...")
        self._escribir_excel(
            nombre_principal, df_export, es_personales_ordenado
        )

        self._reportar("Formateo de Token completado con éxito.")
        return {
            "filas_procesadas": len(df_export),
            "filas_personales": int(es_personales_ordenado.sum()),
        }

    @staticmethod
    def _ordenar_jerarquico(df):
        """Aplica el ordenamiento por Grupo, Tipo de Documento y registros Personales."""

        # 1. Definición de peso para la columna 'Grupo' (Emitidos primero = 0, Recibidos = 1, Otros = 2)
        if "Grupo" in df.columns:
            grupo_clean = (
                df["Grupo"].fillna("").astype(str).str.strip().str.upper()
            )
            condiciones_grupo = [
                grupo_clean.str.contains("EMITIDO"),
                grupo_clean.str.contains("RECIBIDO"),
            ]
            df["_ORDEN_GRUPO"] = np.select(
                condiciones_grupo, [0, 1], default=2
            )
        else:
            df["_ORDEN_GRUPO"] = 0

        # 2. Definición de peso para 'Tipo de Documento' / 'Tipo documento'
        col_tipo_doc = next(
            (
                c
                for c in df.columns
                if c.upper() in ["TIPO DE DOCUMENTO", "TIPO DOCUMENTO"]
            ),
            None,
        )

        if col_tipo_doc:
            doc_clean = (
                df[col_tipo_doc]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
            )
            condiciones_doc = [
                doc_clean.str.contains("FACTURA ELECTRÓNICA|FACTURA ELECTRONICA"),
                doc_clean.str.contains("DOCUMENTO EQUIVALENTE"),
                doc_clean.str.contains("APPLICATION RESPONSE"),
            ]
            # Factura = 0, Doc Equivalente = 1, Application Response = 2, Cualquier otro tipo = 3 (Abajo)
            df["_ORDEN_DOC"] = np.select(condiciones_doc, [0, 1, 2], default=3)
        else:
            df["_ORDEN_DOC"] = 0

        # 3. Aplicar ordenamiento estable
        # CAMBIO CLAVE: Primero ordena por Grupo (_ORDEN_GRUPO), luego envía los personales al final de cada grupo (_ES_PERSONAL)
        sort_cols = ["_ORDEN_GRUPO", "_ES_PERSONAL", "_ORDEN_DOC"]
        ascending_flags = [True, True, True]
        
        if col_tipo_doc:
            sort_cols.append(col_tipo_doc)
            ascending_flags.append(True)

        df = df.sort_values(
            by=sort_cols,
            ascending=ascending_flags,
            kind="stable",
        ).reset_index(drop=True)

        # Limpiar columnas auxiliares
        return df.drop(columns=["_ORDEN_GRUPO", "_ORDEN_DOC"])

    def _cruzar_datos(self, df_principal, df_conta, df_tercer):
        df_principal = df_principal.copy()
        df_conta = df_conta.copy()
        df_tercer = df_tercer.copy()

        # --- VERDADERA SOBRESCRITURA ---
        # Si estas columnas ya existen en la hoja original, las borramos antes de cruzar
        # para que Pandas las recalcule limpiamente sin crear "CONCEPTO_x" o "CONCEPTO_y".
        cols_a_sobrescribir = ['CONCEPTO', 'TERCERO', 'TIPO', 'TIPO-DETALLE', 'NIT_Cruce']
        df_principal = df_principal.drop(columns=[c for c in cols_a_sobrescribir if c in df_principal.columns], errors='ignore')

        # --- LÓGICA CONDICIONAL DE NIT (Emitido vs Recibido) ---
        grupo_clean = df_principal['Grupo'].fillna('').astype(str).str.strip().str.upper()
        
        # Evaluamos el Grupo: Si es EMITIDO toma el Receptor, si no, toma el Emisor.
        df_principal['NIT_Cruce'] = np.where(
            grupo_clean == 'EMITIDO',
            df_principal['NIT Receptor'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip(),
            df_principal['NIT Emisor'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        )

        df_conta['NIT_clean'] = (
            df_conta['NIT'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        )
        df_tercer['NIT_clean'] = (
            df_tercer['NIT'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        )

        df_conta_unica = df_conta.drop_duplicates(subset=['NIT_clean'])

        cols_a_traer = ['NIT_clean']
        if 'TIPO-DETALLE' in df_conta_unica.columns:
            cols_a_traer.append('TIPO-DETALLE')
        if 'TIPO' in df_conta_unica.columns:
            cols_a_traer.append('TIPO')

        # --- CRUCE (Verdadera búsqueda usando NIT_Cruce) ---
        df = df_principal.merge(
            df_conta_unica[cols_a_traer],
            left_on='NIT_Cruce',
            right_on='NIT_clean',
            how='left'
        )

        if 'TIPO-DETALLE' not in df.columns:
            df['TIPO-DETALLE'] = ''
        if 'TIPO' not in df.columns:
            df['TIPO'] = ''
        df['TIPO-DETALLE'] = df['TIPO-DETALLE'].fillna('')
        df['TIPO'] = df['TIPO'].fillna('')

        df['CONCEPTO'] = np.where(
            df['TIPO-DETALLE'].astype(str).str.strip() == '',
            'CONCEPTO NO CREADO',
            df['TIPO-DETALLE'].astype(str).str.strip()
        )

        nits_terceros_unicos = set(df_tercer['NIT_clean'].unique())
        
        # Validamos Terceros basados en el NIT_Cruce también
        df['TERCERO'] = np.where(
            df['NIT_Cruce'].isin(nits_terceros_unicos) & (df['NIT_Cruce'] != ''),
            'CREADO',
            'NO CREADO'
        )
        return df

    @staticmethod
    def _calcular_campos(df):
        prefijo_texto = df['Prefijo'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
        folio_texto = df['Folio'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
        num_ext_base = prefijo_texto + folio_texto

        prefijo_esta_vacio = df['Prefijo'].isna() | (df['Prefijo'].astype(str).str.strip() == '')

        df['Num.Ext'] = np.where(
            prefijo_esta_vacio,
            num_ext_base.str.zfill(10),
            num_ext_base
        )

        df['BASE'] = np.where(df['Total'] == 0, 0, (df['Total'] - df['IVA']))

        es_personales = df['TIPO'].astype(str).str.strip().str.upper() == 'PERSONALES'
        return df, es_personales

    def _escribir_excel(self, nombre_principal, df_export, es_personales):
        try:
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_export.to_excel(writer, index=False, sheet_name=nombre_principal)
                sheet = writer.sheets[nombre_principal]

                red_fill = PatternFill(start_color=RED_FILL_COLOR, end_color=RED_FILL_COLOR, fill_type="solid")
                for row_idx, es_pers in enumerate(es_personales, start=2):
                    if es_pers:
                        for col_idx in range(1, len(df_export.columns) + 1):
                            sheet.cell(row=row_idx, column=col_idx).fill = red_fill

                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                sheet.freeze_panes = 'A2'

                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            pass
                    sheet.column_dimensions[col_letter].width = max(max_len + 2, 10)

                COLUMNAS_OCULTAS = {'TIPO', 'TIPO-DETALLE', 'BASE', 'Num.Ext'}
                for idx, col_name in enumerate(df_export.columns, start=1):
                    if col_name in COLUMNAS_OCULTAS:
                        sheet.column_dimensions[get_column_letter(idx)].hidden = True
        except (ErrorUsuario, ErrorSistema):
            raise
        except Exception as e:
            raise ErrorSistema(f"Error escribiendo la hoja Token: {e}") from e