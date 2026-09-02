import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from errors import ErrorSistema, ErrorUsuario, logger

TIPOS_DOC_VALIDOS = {"Factura electrónica", "Nota credito"}
GRUPOS_VALIDOS = {"Emitido", "Recibido"}

COLUMNA_GRUPO = "Grupo"
COLUMNA_TIPO_DOC = "Tipo de documento"
COLUMNA_IVA = "IVA"
COLUMNA_FECHA = "Fecha Emisión"

HOJA_INFORME_IVA = "Informe IVA"

MESES_ESPANOL = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

# informes_iva.py — imports (agregar)
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape as escapar_xml

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
class GeneradorInformeIVA:

# informes_iva.py — GeneradorInformeIVA.__init__ (reemplazar)
    def __init__(self, ruta_archivo, hoja_token, progress_callback=None):
        self.ruta_archivo = ruta_archivo
        self.hoja_token = hoja_token
        self.progress_callback = progress_callback or (lambda msg: None)
        self.df_filtrado = None
        self.resumen = None
        self.periodo_texto = ""

    def _log(self, mensaje):
        logger.info(mensaje)
        self.progress_callback(mensaje)

    def generar(self):
            self._log(f"Leyendo hoja de token: {self.hoja_token}")
            df = self._cargar_hoja(self.ruta_archivo, self.hoja_token)

            for col in (COLUMNA_GRUPO, COLUMNA_TIPO_DOC, COLUMNA_IVA):
                if col not in df.columns:
                    raise ErrorUsuario(
                        f"La hoja '{self.hoja_token}' no contiene la columna requerida '{col}'."
                    )

            df[COLUMNA_GRUPO] = df[COLUMNA_GRUPO].astype(str).str.strip()
            df[COLUMNA_TIPO_DOC] = df[COLUMNA_TIPO_DOC].astype(str).str.strip()
            df[COLUMNA_IVA] = (
                pd.to_numeric(df[COLUMNA_IVA], errors="coerce").fillna(0)
            )

            # Construir la condición base de filtrado
            condicion_filtro = (
                df[COLUMNA_TIPO_DOC].isin(TIPOS_DOC_VALIDOS)
                & df[COLUMNA_GRUPO].isin(GRUPOS_VALIDOS)
            )

            # NUEVO: Excluir los registros donde CONCEPTO indique "PERSONALES"
            if "CONCEPTO" in df.columns:
                # Usamos upper() y strip() para evitar problemas con espacios o minúsculas/mayúsculas
                mascara_personales = df["CONCEPTO"].astype(str).str.strip().str.upper() == "PERSONALES"
                condicion_filtro = condicion_filtro & ~mascara_personales
                
                registros_omitidos = mascara_personales.sum()
                if registros_omitidos > 0:
                    self._log(f"Se ignoraron {registros_omitidos} registro(s) por tener el concepto 'PERSONALES'.")

            df_filtrado = df[condicion_filtro].copy()

            if df_filtrado.empty:
                raise ErrorUsuario(
                    "No se encontraron registros que cumplan con los criterios indicados, "
                    "o todos fueron filtrados/excluidos."
                )

            self._log(f"Registros filtrados: {len(df_filtrado)}")

            df_filtrado["Signo"] = df_filtrado.apply(self._calcular_signo, axis=1)
            df_filtrado["IVA Ajustado"] = (
                df_filtrado[COLUMNA_IVA] * df_filtrado["Signo"]
            )

            # Determinar el período analizando las fechas de emisión
            periodo_texto = self._obtener_texto_periodo(df)

            resumen = self._construir_resumen(df_filtrado)

            self.df_filtrado = df_filtrado
            self.resumen = resumen
            self.periodo_texto = periodo_texto

            self._log("Escribiendo informe en el archivo de Excel...")
            self._escribir_informe(df_filtrado, resumen, periodo_texto)

            self._log("Informe de IVA generado correctamente.")
            return resumen

    @staticmethod
    def _obtener_texto_periodo(df):
        """Busca la columna de Fecha Emisión y construye el texto del período."""
        col_fecha = next(
            (
                c
                for c in df.columns
                if c.strip().lower()
                in ["fecha emisión", "fecha emision", "fecha_emision"]
            ),
            None,
        )

        if not col_fecha:
            return ""

        fechas = pd.to_datetime(
            df[col_fecha], errors="coerce", dayfirst=True
        ).dropna()
        if fechas.empty:
            return ""

        meses = sorted(fechas.dt.month.unique())
        anios = sorted(fechas.dt.year.unique())

        nombres_meses = [MESES_ESPANOL[m] for m in meses if m in MESES_ESPANOL]

        if not nombres_meses:
            return ""

        if len(nombres_meses) == 1:
            texto_meses = nombres_meses[0]
        else:
            texto_meses = f"{nombres_meses[0]} - {nombres_meses[-1]}"

        texto_anios = (
            str(anios[0])
            if len(anios) == 1
            else f"{anios[0]}-{anios[-1]}"
        )
        return f"PERÍODO: {texto_meses.upper()} {texto_anios}"

    @staticmethod
    def _calcular_signo(fila):
        grupo = fila[COLUMNA_GRUPO]
        tipo_doc = fila[COLUMNA_TIPO_DOC]

        if grupo == "Emitido":
            return 1 if tipo_doc == "Factura electrónica" else -1
        if grupo == "Recibido":
            return -1 if tipo_doc == "Factura electrónica" else 1
        return 0
    @staticmethod
    def _seccion_pdf(texto, estilo, ancho):
        tabla = Table([[Paragraph(texto, estilo)]], colWidths=[ancho])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return tabla

    @staticmethod
    def _cargar_hoja(ruta, hoja):
        try:
            return pd.read_excel(ruta, sheet_name=hoja)
        except Exception as e:
            logger.error(f"No se pudo leer la hoja '{hoja}': {e}")
            raise ErrorSistema(
                f"No se pudo leer la hoja '{hoja}' del archivo: {e}"
            )
    @staticmethod
    def _texto_celda(valor):
        if pd.isna(valor):
            return ""
        if hasattr(valor, "strftime"):
            return valor.strftime("%d/%m/%Y")
        return escapar_xml(str(valor))
    # informes_iva.py — nuevos métodos (agregar a la clase GeneradorInformeIVA)
    def generar_pdf(self, ruta_pdf=None):
            if self.df_filtrado is None or self.resumen is None:
                raise ErrorSistema(
                    "Debes generar el informe de IVA antes de exportarlo a PDF."
                )

            if ruta_pdf is None:
                # Buscar el primer Nombre Emisor donde el Grupo sea Emitido
                nombre_part1 = ""
                if "Grupo" in self.df_filtrado.columns and "Nombre Emisor" in self.df_filtrado.columns:
                    df_emitidos = self.df_filtrado[self.df_filtrado["Grupo"].astype(str).str.strip() == "Emitido"]
                    if not df_emitidos.empty:
                        val_emisor = df_emitidos["Nombre Emisor"].iloc[0]
                        if pd.notna(val_emisor) and str(val_emisor).strip():
                            nombre_part1 = str(val_emisor).strip() + " - "

                # Extraer periodo limpio o usar el texto base
                periodo_limpio = self.periodo_texto.replace("PERÍODO:", "").strip() if self.periodo_texto else ""
                sufijo_periodo = f" - {periodo_limpio}" if periodo_limpio else ""

                nombre_sugerido = f"{nombre_part1}Informe de IVA{sufijo_periodo}.pdf"
                
                # Limpiar caracteres inválidos para nombres de archivos en Windows/Linux
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    nombre_sugerido = nombre_sugerido.replace(char, '')

                ruta_pdf = str(Path(self.ruta_archivo).with_name(nombre_sugerido))

            self._log("Generando documento PDF...")


            margen = 0.8 * cm
            ancho_pagina, _ = landscape(letter)
            ancho_util = ancho_pagina - 2 * margen

            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle(
                "TituloInforme", parent=styles["Title"],
                fontName="Helvetica-Bold", fontSize=14, textColor=colors.black,
                alignment=TA_CENTER, spaceAfter=2,
            )
            estilo_subtitulo = ParagraphStyle(
                "Subtitulo", parent=styles["Normal"],
                fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#333333"),
                alignment=TA_CENTER, spaceAfter=0,
            )
            estilo_seccion = ParagraphStyle(
                "Seccion", parent=styles["Heading2"],
                fontName="Helvetica-Bold", fontSize=10, textColor=colors.black,
            )

            # FUENTE DE DETALLE
            estilo_celda = ParagraphStyle(
                "Celda", parent=styles["Normal"],
                fontName="Helvetica", fontSize=5.5, leading=6.5, textColor=colors.black,
            )
            estilo_celda_num = ParagraphStyle(
                "CeldaNum", parent=estilo_celda, alignment=TA_RIGHT,
            )
            estilo_encabezado_tabla = ParagraphStyle(
                "EncTabla", parent=styles["Normal"],
                fontName="Helvetica-Bold", fontSize=6, leading=7,
                textColor=colors.white, alignment=TA_CENTER,
            )

            estilo_celda_resumen = ParagraphStyle(
                "CeldaResumen", parent=styles["Normal"],
                fontName="Helvetica", fontSize=8, leading=10, textColor=colors.black,
            )
            estilo_celda_resumen_num = ParagraphStyle(
                "CeldaResumenNum", parent=estilo_celda_resumen, alignment=TA_RIGHT,
            )
            estilo_encabezado_resumen = ParagraphStyle(
                "EncResumen", parent=styles["Normal"],
                fontName="Helvetica-Bold", fontSize=8, leading=10,
                textColor=colors.white, alignment=TA_CENTER,
            )

            historia = []
            historia.append(Paragraph("INFORME DE IVA", estilo_titulo))
            if self.periodo_texto:
                historia.append(Paragraph(escapar_xml(self.periodo_texto), estilo_subtitulo))
                historia.append(Spacer(1, 2))
                
            historia.append(Paragraph(
                f"Hoja origen: {escapar_xml(self.hoja_token)}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                estilo_subtitulo,
            ))
            historia.append(Spacer(1, 4))
            historia.append(HRFlowable(width="100%", thickness=1, color=colors.black,
                                        spaceBefore=2, spaceAfter=8))

            # --- SECCIÓN 1: RESUMEN GENERAL ---
            historia.append(self._seccion_pdf("RESUMEN GENERAL", estilo_seccion, ancho_util))
            historia.append(Spacer(1, 4))

            encabezados_resumen = [
                "Concepto", "Facturas<br/>(cant. / IVA)", "Notas Crédito<br/>(cant. / IVA)", "IVA Neto",
            ]
            datos_resumen = [[Paragraph(h, estilo_encabezado_resumen) for h in encabezados_resumen]]

            nombres_conceptos = {
                "Emitido": "IVA DESCONTABLE",
                "Recibido": "IVA GENERADO",
            }
            for grupo in ("Emitido", "Recibido"):
                d = self.resumen["grupos"][grupo]
                datos_resumen.append([
                    Paragraph(nombres_conceptos[grupo], estilo_celda_resumen),
                    Paragraph(f"{d['facturas_cantidad']} / {self._fmt(d['facturas_iva'])}", estilo_celda_resumen_num),
                    Paragraph(f"{d['notas_cantidad']} / {self._fmt(d['notas_iva'])}", estilo_celda_resumen_num),
                    Paragraph(self._fmt(d["iva_neto"]), estilo_celda_resumen_num),
                ])

            t = self.resumen["totales"]
            datos_resumen.append([
                Paragraph("<b>TOTAL</b>", estilo_celda_resumen),
                "", "",
                Paragraph(f"<b>{self._fmt(t['iva_a_pagar'])}</b>", estilo_celda_resumen_num),
            ])

            tabla_resumen = Table(
                datos_resumen,
                colWidths=[ancho_util * 0.34, ancho_util * 0.24, ancho_util * 0.24, ancho_util * 0.18],
                hAlign="LEFT",
            )
            tabla_resumen.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F2F2")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DDDDDD")),
                ("SPAN", (1, -1), (2, -1)),
                ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ]))
            historia.append(tabla_resumen)
            historia.append(Spacer(1, 10))

    # --- SECCIÓN 2: DETALLE DE REGISTROS ---
            historia.append(self._seccion_pdf("DETALLE DE REGISTROS", estilo_seccion, ancho_util))
            historia.append(Spacer(1, 4))

            columnas_excluidas_pdf = {
                "CUFE/CUDE", "CONCEPTO", "TERCERO", "TIPO", 
                "TIPO-DETALLE", "BASE", "Num.Ext", "Signo", "IVA Ajustado","ICA", "IC", "INC", "Timbre", "INC Bolsas", "IN Carbono", 
                "IN Combustibles", "IC Datos", "ICL", "INPP", "IBUA", "ICUI", "Rete IVA", "Rete Renta", "Rete ICA","IVA Ajustado"
            }

            columnas = [
                c for c in self.df_filtrado.columns 
                if str(c).strip() not in columnas_excluidas_pdf
            ]

            anchos = self._calcular_anchos_columnas(self.df_filtrado, columnas, ancho_util)

            filas_detalle = [[Paragraph(escapar_xml(str(c)), estilo_encabezado_tabla) for c in columnas]]

            # Columnas que van como moneda con decimales
            columnas_moneda = {COLUMNA_IVA, "Total"}
            
            # Columnas que son identificadores numéricos (NIT, etc.) pero deben mostrarse sin decimales .0
            columnas_identificadores = {"NIT Emisor", "NIT Receptor", "Forma de Pago", "Medio de Pago"}

            for _, registro in self.df_filtrado.iterrows():
                fila = []
                for col in columnas:
                    valor = registro[col]
                    
                    if col in columnas_moneda:
                        fila.append(Paragraph(self._fmt(valor), estilo_celda_num))
                    elif col in columnas_identificadores:
                        fila.append(Paragraph(self._fmt_entero(valor), estilo_celda_num if "NIT" in col else estilo_celda))
                    else:
                        fila.append(Paragraph(self._texto_celda(valor), estilo_celda))
                filas_detalle.append(fila)

            tabla_detalle = Table(filas_detalle, colWidths=anchos, repeatRows=1, hAlign="LEFT")
            tabla_detalle.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0B0B0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 1.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
            ]))
            historia.append(tabla_detalle)

            doc = SimpleDocTemplate(
                ruta_pdf, pagesize=landscape(letter),
                leftMargin=margen, rightMargin=margen,
                topMargin=1.0 * cm, bottomMargin=1.2 * cm,
                title="Informe de IVA",
            )

            def _pie_pagina(canvas_obj, doc_obj):
                canvas_obj.saveState()
                canvas_obj.setStrokeColor(colors.HexColor("#999999"))
                canvas_obj.line(margen, 0.9 * cm, ancho_pagina - margen, 0.9 * cm)
                canvas_obj.setFont("Helvetica", 7)
                canvas_obj.setFillColor(colors.HexColor("#555555"))
                canvas_obj.drawString(margen, 0.6 * cm,
                                    f"Informe generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                canvas_obj.drawRightString(ancho_pagina - margen, 0.6 * cm, f"Página {doc_obj.page}")
                canvas_obj.restoreState()

            try:
                doc.build(historia, onFirstPage=_pie_pagina, onLaterPages=_pie_pagina)
            except Exception as e:
                raise ErrorSistema(f"No se pudo generar el PDF del informe: {e}")

            self._log("PDF generado correctamente.")
            return ruta_pdf

    @staticmethod
    def _fmt(valor):
        try:
            return f"$ {float(valor):,.2f}"
        except (TypeError, ValueError):
            return escapar_xml(str(valor))
        
    @staticmethod
    def _fmt_entero(valor):
        """Limpia los números de identificación para que no muestren decimales '.0'."""
        if pd.isna(valor):
            return ""
        try:
            val_float = float(valor)
            if val_float.is_integer():
                return str(int(val_float))
            return str(valor)
        except (TypeError, ValueError):
            return escapar_xml(str(valor))

    @staticmethod
    def _calcular_anchos_columnas(df, columnas, ancho_disponible):
        longitudes = []
        for col in columnas:
            peso_extra = 1.5 if col in {COLUMNA_IVA, "Total", "NIT Emisor", "NIT Receptor"} else 1.0
            
            valores = [
                "" if pd.isna(v) else str(v)
                for v in df[col].tolist()
            ]
            max_len = max([len(str(col))] + [len(v) for v in valores])
            longitudes.append(max(max_len * peso_extra, 4))
            
        total = sum(longitudes)
        return [max(ancho_disponible * (l / total), 15) for l in longitudes]

    def _construir_resumen(self, df):
        resumen = {"grupos": {}, "totales": {}}

        for grupo in ("Emitido", "Recibido"):
            df_grupo = df[df[COLUMNA_GRUPO] == grupo]
            facturas = df_grupo[
                df_grupo[COLUMNA_TIPO_DOC] == "Factura electrónica"
            ]
            notas = df_grupo[df_grupo[COLUMNA_TIPO_DOC] == "Nota credito"]

            resumen["grupos"][grupo] = {
                "facturas_cantidad": len(facturas),
                "facturas_iva": round(facturas[COLUMNA_IVA].sum(), 2),
                "notas_cantidad": len(notas),
                "notas_iva": round(notas[COLUMNA_IVA].sum(), 2),
                "iva_neto": round(df_grupo["IVA Ajustado"].sum(), 2),
            }

        iva_emitido = resumen["grupos"]["Emitido"]["iva_neto"]
        iva_recibido = resumen["grupos"]["Recibido"]["iva_neto"]

        resumen["totales"] = {
            "iva_emitido": iva_emitido,
            "iva_recibido": iva_recibido,
            "iva_a_pagar": round(iva_emitido + iva_recibido, 2),
            "total_registros": len(df),
        }
        return resumen

    def _escribir_informe(self, df, resumen, periodo_texto=""):
        try:
            libro = load_workbook(self.ruta_archivo)
        except Exception as e:
            raise ErrorSistema(
                f"No se pudo abrir el archivo para escribir el informe: {e}"
            )

        if HOJA_INFORME_IVA in libro.sheetnames:
            del libro[HOJA_INFORME_IVA]
        hoja = libro.create_sheet(HOJA_INFORME_IVA, 0)

        estilo_titulo = Font(size=14, bold=True, color="1F2937")
        estilo_periodo = Font(size=11, bold=True, color="2563EB")
        estilo_encabezado_seccion = Font(size=11, bold=True, color="FFFFFF")
        relleno_seccion = PatternFill("solid", fgColor="2563EB")
        estilo_encabezado_tabla = Font(bold=True, color="FFFFFF")
        relleno_encabezado_tabla = PatternFill("solid", fgColor="374151")
        relleno_total = PatternFill("solid", fgColor="D1FAE5")
        estilo_total = Font(bold=True)
        borde_fino = Border(*(Side(style="thin", color="D3D8E0"),) * 4)
        formato_moneda = "#,##0.00"
        relleno_ventas = PatternFill("solid", fgColor="EDFFC3")
        relleno_compras = PatternFill("solid", fgColor="E0E7FF")

        fila = 1
        hoja.cell(row=fila, column=1, value="INFORME DE IVA").font = estilo_titulo

        if periodo_texto:
            celda_periodo = hoja.cell(row=fila, column=3, value=periodo_texto)
            celda_periodo.font = estilo_periodo

        fila += 1
        hoja.cell(
            row=fila, column=1, value=f"Hoja origen: {self.hoja_token}"
        ).font = Font(italic=True, color="5B6472")
        fila += 2

        hoja.cell(row=fila, column=1, value="RESUMEN GENERAL")
        hoja.cell(row=fila, column=1).font = estilo_encabezado_seccion
        hoja.merge_cells(
            start_row=fila, start_column=1, end_row=fila, end_column=4
        )
        for c in range(1, 5):
            hoja.cell(row=fila, column=c).fill = relleno_seccion
        fila += 1

        encabezados_resumen = [
            "Concepto",
            "Facturas (cant. / IVA)",
            "Notas Crédito (cant. / IVA)",
            "IVA Neto",
        ]
        for c, texto in enumerate(encabezados_resumen, start=1):
            celda = hoja.cell(row=fila, column=c, value=texto)
            celda.font = estilo_encabezado_tabla
            celda.fill = relleno_encabezado_tabla
            celda.border = borde_fino
        fila += 1

        nombres_conceptos = {
            "Emitido": "IVA DEVOLUCION DE VENTAS",
            "Recibido": "IVA DEVOLUCION DE COMPRAS",
        }
        for grupo in ("Emitido", "Recibido"):
            datos = resumen["grupos"][grupo]
            hoja.cell(
                row=fila, column=1, value=nombres_conceptos[grupo]
            ).border = borde_fino
            hoja.cell(
                row=fila,
                column=2,
                value=f"{datos['facturas_cantidad']} / {datos['facturas_iva']:,.2f}",
            ).border = borde_fino
            hoja.cell(
                row=fila,
                column=3,
                value=f"{datos['notas_cantidad']} / {datos['notas_iva']:,.2f}",
            ).border = borde_fino
            c4 = hoja.cell(row=fila, column=4, value=datos["iva_neto"])
            c4.number_format = formato_moneda
            c4.border = borde_fino
            fila += 1

        totales = resumen["totales"]
        hoja.cell(row=fila, column=1, value="IVA GENERADO").font = estilo_total
        hoja.cell(row=fila, column=1).fill = relleno_total
        celda_total = hoja.cell(
            row=fila, column=4, value=totales["iva_a_pagar"]
        )
        celda_total.font = estilo_total
        celda_total.fill = relleno_total
        celda_total.number_format = formato_moneda
        for c in range(1, 5):
            hoja.cell(row=fila, column=c).border = borde_fino
        fila += 3

        hoja.cell(row=fila, column=1, value="DETALLE DE REGISTROS")
        hoja.cell(row=fila, column=1).font = estilo_encabezado_seccion
        hoja.merge_cells(
            start_row=fila,
            start_column=1,
            end_row=fila,
            end_column=len(df.columns),
        )
        for c in range(1, len(df.columns) + 1):
            hoja.cell(row=fila, column=c).fill = relleno_seccion
        fila += 1

        columnas_detalle = list(df.columns)
        for c, nombre_col in enumerate(columnas_detalle, start=1):
            celda = hoja.cell(row=fila, column=c, value=nombre_col)
            celda.font = estilo_encabezado_tabla
            celda.fill = relleno_encabezado_tabla
            celda.border = borde_fino
        fila_inicio_detalle = fila
        fila += 1

        for _, registro in df.iterrows():
            if registro[COLUMNA_GRUPO] == "Emitido":
                relleno_fila = relleno_ventas
            elif registro[COLUMNA_GRUPO] == "Recibido":
                relleno_fila = relleno_compras
            else:
                relleno_fila = None

            for c, nombre_col in enumerate(columnas_detalle, start=1):
                celda = hoja.cell(
                    row=fila, column=c, value=registro[nombre_col]
                )
                celda.border = borde_fino
                if relleno_fila:
                    celda.fill = relleno_fila

                if nombre_col in (COLUMNA_IVA, "IVA Ajustado"):
                    celda.number_format = formato_moneda
            fila += 1

        for c, nombre_col in enumerate(columnas_detalle, start=1):
            max_len = max(
                [len(str(nombre_col))]
                + [len(str(v)) for v in df[nombre_col].astype(str).tolist()]
            )
            hoja.column_dimensions[get_column_letter(c)].width = min(
                max(max_len + 2, 12), 45
            )

        hoja.freeze_panes = hoja.cell(row=fila_inicio_detalle + 1, column=1)

        try:
            libro.save(self.ruta_archivo)
        except Exception as e:
            raise ErrorSistema(
                f"No se pudo guardar el informe en el archivo: {e}"
            )